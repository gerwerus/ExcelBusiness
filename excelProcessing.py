from typing import Iterable
from PySide6.QtCore import QThread, QDate
from openpyxl import load_workbook

from consts import cellColumn


class ExcelTread(QThread):
    def __init__(self, mainMenu):
        super().__init__()
        self.mainMenu = mainMenu
        self.objects = 0
        self.ln = 0

    def setObjects(self, objects: Iterable, ln=0):
        self.objects = objects
        self.ln = ln

    def run(self):
        try:
            formDate: QDate = self.mainMenu.dateEdit.date()
            for index, obj in enumerate(self.objects):
                startDate = QDate(
                    int(obj.getDateStart()), 9, 1
                )  # год из таблицы, месяц 09, день 01
                if startDate >= formDate:
                    return
                dt = startDate.daysTo(formDate)

                view = list(obj.getView())
                dates = obj.getPracticeDates()
                for ind, el in enumerate(view):
                    currentStartYear = startDate.year() + (el[0] + 1) // 2 - 1
                    if el[1] == "Научно-исследовательская работа":
                        continue
                    if (
                        QDate(currentStartYear, 9, 1)
                        <= formDate
                        <= QDate(currentStartYear + 1, 8, 31)
                    ):  # Берем практики за год
                        currentDates = dates.get(str(currentStartYear), None)
                        # print(currentDates, el[0], el[1])
                        currentPracticeDate = ["", ""]
                        if currentDates:
                            currentPracticeDate = currentDates[
                                ExcelProccessing.translateVal(el[1])
                            ]
                        self.mainMenu.addLine(
                            [
                                obj.getInstituteName(),
                                obj.getCode(),
                                obj.getProfile(),
                                el[0],
                                el[1],
                                el[2],
                                el[3],
                                currentPracticeDate[0].toString("dd.MM.yyyy"),
                                currentPracticeDate[1].toString("dd.MM.yyyy"),
                                "; ".join(el[4]),
                                str(formDate.year()),
                            ]
                        )
                obj.book.close()
                self.mainMenu.changeProgressSignal.emit(
                    int(((index + 1) / self.ln) * 100)
                )
        finally:
            self.mainMenu.getFilesButton.setEnabled(True)
            self.mainMenu.getFilesFromDirButton.setEnabled(True)


class ExcelProccessing:
    def __init__(self, filename):
        self.filename = filename
        self.book = load_workbook(filename)
        self.titleList = self.book["Титул"]
        self.practiceList = self.book["Практики"]
        self.planList = self.book["План"]
        self.graphList = self.book["График"]

    def getInstituteName(self) -> str:
        return self.titleList["D38"].value.replace(
            "Институт ", ""
        )  # Жесткая привязка к D38

    def getCode(self) -> str:
        return self.titleList["D29"].value

    def getProfile(self) -> str:
        val = self.titleList["D30"].value
        if val:
            return val
        return "Отсутсвует"

    def getDateStart(self) -> str:
        return self.titleList["W40"].value

    def getView(self) -> zip:
        practiceView = []
        practiceTyp = []
        practiceSem = []
        double = 0

        nameCourseColumn = nameSemColumn = nameViewColumn = ""
        for i in range(1, 5):
            for name in cellColumn:
                if self.practiceList[name + str(i)].value == "Курс":
                    nameCourseColumn = name
                    continue
                if self.practiceList[name + str(i)].value == "Сем. курса":
                    nameSemColumn = name
                    continue
                if self.practiceList[name + str(i)].value == "Название практики":
                    nameViewColumn = name
                    continue
            if not (nameSemColumn):
                nameSemColumn = "no info"  # Если отсутствует колонка семестра
            if nameCourseColumn and nameSemColumn and nameViewColumn:
                break
        else:
            raise Exception("Не удалось распарсить лист Практик")
        for i in range(3, self.practiceList.max_row + 1):
            view = self.practiceList[nameViewColumn + str(i)].value
            typ = self.practiceList[
                cellColumn[cellColumn.index(nameViewColumn) + 1] + str(i)
            ].value
            course = self.practiceList[nameCourseColumn + str(i)].value
            if nameSemColumn == "no info":
                sem = 2
            else:
                sem = self.practiceList[nameSemColumn + str(i)].value
            if view:
                if "Вид" in view:
                    practiceView.append(view[14:].rstrip())  # Убирает "Вид практики: "
                    double = 0
            if typ:
                practiceTyp.append(typ)
                double += 1
                if double > 1:
                    practiceView.append(practiceView[-1])
            if course:
                practiceSem.append((int(course) - 1) * 2 + sem)
        practiceComp = []
        zeColumn = "Неопределено"
        nameColumn = "Неопределено"
        hoursColumn = "Неопределено"
        for col in cellColumn:
            if (
                self.planList[col + "2"].value == "Наименование"
                or self.planList[col + "3"].value == "Наименование"
                or self.planList[col + "4"].value == "Наименование"
            ):
                nameColumn = col
            elif (
                self.planList[col + "1"].value == "з.е."
                or self.planList[col + "2"].value == "з.е."
                or self.planList[col + "3"].value == "з.е."
            ):  # Брать экспертное или фактическое
                zeColumn = col
            elif (
                self.planList[col + "1"].value == "Итого акад.часов"
                or self.planList[col + "2"].value == "Итого акад.часов"
                or self.planList[col + "3"].value == "Итого акад.часов"
            ):
                hoursColumn = col
            if (
                zeColumn != "Неопределено"
                and hoursColumn != "Неопределено"
                and nameColumn != "Неопределено"
            ):
                break
        else:
            raise Exception("Не удалось распарсить колонки")
        laborTime = []
        for i in range(len(practiceView)):
            startIndex = 1
            while self.planList[nameColumn + str(startIndex)].value != practiceTyp[i]:
                startIndex += 1
                if startIndex > self.planList.max_row:
                    break
            laborTime.append(
                str(self.planList[hoursColumn + str(startIndex)].value)
                + "/"
                + str(self.planList[zeColumn + str(startIndex)].value)
            )
            compList = self.book["Компетенции"]
            currentCode = ""
            currentText = ""
            allComp = []
            for row in range(2, compList.max_row + 1):
                code = compList["B" + str(row)].value
                comp = compList["D" + str(row)].value
                if code:
                    currentCode = code
                    currentText = comp
                if comp == practiceTyp[i]:
                    if not (currentText):
                        currentText = "Не указано"  # Файлы с пустыми компетенциями
                    allComp.append(currentCode + " - " + currentText)
            practiceComp.append(allComp)  # practiceComp - список компетенций
        return zip(practiceSem, practiceView, practiceTyp, laborTime, practiceComp)

    def getPracticeDates(self) -> dict:
        # print(self.filename)
        startCol = "B"
        startRow = 4
        diapazon = 6
        secondDiapazon = 9
        yearStep = 17
        startYear = self.getDateStart()
        yearsCounter = 0
        for row in range(2, self.graphList.max_row):
            val = self.graphList["A" + str(row)].value
            if not (val):
                continue
            if "учебный график" in val:
                yearsCounter += 1
        allDates = {}
        for year in range(yearsCounter):
            currentMonth = 9
            currentYear = int(startYear) + year
            currentStartRow = startRow + year * yearStep
            currentDay = 0
            practiceDates = {"У": [], "П": [], "Пд": []}
            for cell in cellColumn[cellColumn.index(startCol) :]:
                checkNone = False
                for row in range(
                    currentStartRow, currentStartRow + diapazon
                ):  # Воскресенье не берем
                    prevDay = currentDay
                    currentDay = self.graphList[cell + str(row)].value
                    if not (currentDay):
                        continue
                    currentDay = int(currentDay)
                    if prevDay:
                        if prevDay > currentDay:
                            currentMonth = (currentMonth + 1) % 13
                            if currentMonth == 0:
                                currentMonth = 1
                                currentYear += 1
                    currentDate = QDate(currentYear, currentMonth, currentDay)
                    valCell = self.graphList[cell + str(row + secondDiapazon)]
                    val = valCell.value
                    if valCell.fill.start_color.index == 16:
                        val = "У"
                    if val in ["У", "П", "Пд", "П Н"]:
                        if val == "П Н":
                            val = "П"
                        ln = len(practiceDates[val])
                        if ln == 0:
                            practiceDates[val].append(currentDate)
                        elif ln == 1:
                            practiceDates[val].append(currentDate)
                            checkNone = val
                        elif ln == 2:
                            practiceDates[val][1] = currentDate
                            checkNone = val
                    elif val is not None:
                        checkNone = False
                    elif checkNone:
                        practiceDates[checkNone][1] = currentDate
                    # if cell == "AB" and currentYear == 2023:
                    #     print(val, currentDate)
            allDates.update({str(int(startYear) + year): practiceDates})
        return allDates

    @staticmethod
    def translateVal(val: str) -> str:
        if val == "Учебная практика":
            return "У"
        if val == "Производственная практика":
            return "П"
        if val == "Преддипломная практика":
            return "Пд"
        raise ValueError(f"{val} is not Practice type!")
